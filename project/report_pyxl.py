# -*- coding: utf-8 -*-

"""
@author Jacky
@desc report operate module
@date 2020/08/18
"""

import os
import openpyxl
from openpyxl.chart import Series, ScatterChart, Reference
from openpyxl.styles import Color, Font, Alignment, Border, Side
import gloabl_var as gl


class ReportUtil(object):
    """

    """

    def __init__(self):
        self.__wb = ''
        self.__ws = ''
        self.__report_save_path = ''
        self.__report_demo_path = ''

    def generate_report(self, target_dir):
        """generate report file

        :param target_dir: To store report file
        :return report_file: dst report file path
        """
        print('Enter', ReportUtil.__name__, 'generate_report')

        report = gl.get_value('report_name')
        self.__report_demo_path = os.path.join('report', report)
        self.__report_save_path = os.path.join(target_dir, report)

        print('generate report done!')

    def open_workbook(self):
        """open report file and return file handle

        :param report_file:
        :return:new_excel
        """
        print('Enter', ReportUtil.__name__, 'open_workbook ......')
        self.__wb = openpyxl.load_workbook(self.__report_demo_path)

        print('open success!')

    def save_workbook(self):
        """

        :param report_file:
        :return:
        """
        self.__wb.save(self.__report_save_path)
        print('Enter', ReportUtil.__name__, 'save_workbook:', self.__report_save_path)

    def write_report(self, camera, test_item, data, sub_item=None):
        """

        :param sub_item:
        :param camera: specified Front or Main camera
        :param test_item: For example defect,AWB,DR etc.
        :param data: objective data
        :param : sub_item
        :return:
        """

        if 'TE255' == test_item:  # For defect
            self.write_defect_data(camera, data, sub_item)
        elif 'OB' == test_item:
            self.write_ob_data(camera, data)
        elif 'IMAGE_SIZE' == test_item:
            self.write_size_data(camera, data)
        elif 'POWER_LINE' == test_item:
            self.write_power_line_data(camera, data)
        elif 'COLOR_UNIFORM' == test_item:
            self.write_color_uniform_data(camera, data, sub_item)
        elif 'LUMA_UNIFORM' == test_item:
            self.write_luma_uniform_data(camera, data)
        elif 'COLOR_ACCURACY' == test_item:
            self.write_color_accuracy_data(camera, data, sub_item)
        elif 'SATURATION' == test_item:
            self.write_saturation_data(camera, data, sub_item)
        elif 'WB' == test_item:
            self.write_wb_data(camera, data, sub_item)
        elif 'FREE' == test_item:  # color checker free test
            self.write_free_data(camera, data, sub_item)
        elif 'CORNER' == test_item:
            self.write_corner_data(camera, data, sub_item)
        elif 'DR' == test_item:
            self.write_dr_data(camera, data)

        print(ReportUtil.__name__, 'write_report:',
              ' camera:', camera,
              ' test item:', test_item,
              ' sub item:', sub_item,
              'Done'
              )

    def write_defect_data(self, camera, data, sub_item):
        """

        :param camera:
        :param data:
        :param sub_item:
        :return:
        """
        ws = self.__wb.get_sheet_by_name('坏点及视场角')
        if 'dark' == sub_item:
            if 'front' == camera:
                ws['D19'] = data[0]
                ws['D20'] = data[1]
            else:  # for main camera
                ws['E19'] = data[0]
                ws['E20'] = data[1]
        else:
            if 'front' == camera:
                ws['D15'] = data[0]
                ws['D16'] = data[1]
            else:  # for main camera
                ws['E15'] = data[0]
                ws['E16'] = data[1]

    def write_ob_data(self, camera, data):
        """

        :param camera:
        :param data:
        :return:
        """
        ws = self.__wb.get_sheet_by_name('暗电流')
        if 'front' == camera:
            ws['D6'] = data[3][data[0]]  # R
            ws['E6'] = data[2][data[0]]  # G
            ws['F6'] = data[1][data[0]]  # B
        else:  # for main camera
            ws['D7'] = data[3][data[0]]  # R
            ws['E7'] = data[2][data[0]]  # G
            ws['F7'] = data[1][data[0]]  # B

        row_start_index = 12
        files = data[4]
        image_num = len(files)
        row_index = row_start_index
        for index in range(image_num):
            cnt_cell = 'B' + str(row_index)
            ws[cnt_cell] = index + 1
            # file_cell = 'B' + str(start_index)
            # ws[file_cell] = files[index]
            r_cell = 'C' + str(row_index)
            g_cell = 'D' + str(row_index)
            b_cell = 'E' + str(row_index)
            ws[r_cell] = data[3][index]  # R
            ws[g_cell] = data[2][index]  # G
            ws[b_cell] = data[1][index]  # B
            row_index += 1

        chart = ScatterChart()  # LineChart()
        chart.title = "暗电流"  #图的标题
        # chart.style = 2  # 线条的style
        chart.y_axis.title = '灰度值'  # y坐标的标题
        # c1.x_axis.title = "Date"  # x坐标的标题

        x_value = Reference(ws, min_col=2, min_row=row_start_index, max_row=(row_start_index + image_num - 1))
        line_color = ['FF0000', '00FF00', '0000FF']
        for i in range(3, 6):
            value = Reference(ws, min_col=i, min_row=row_start_index - 1, max_row=(row_start_index + image_num - 1))
            series = Series(value, x_value, title_from_data=True)
            series.graphicalProperties.line.solidFill = line_color[i-3]
            chart.series.append(series)

        ws.add_chart(chart, "B30")  # 将图表添加到 sheet中

    def write_size_data(self, camera, data):
        """

        :param camera:
        :param data:
        :return:
        """
        ws = self.__wb.get_sheet_by_name('像素及分辨率')
        if 'front' == camera:
            ws['D12'] = data[1]
            ws['E12'] = data[0]
        else:  # for main camera
            ws['D13'] = data[1]
            ws['E13'] = data[0]

    def write_power_line_data(self, camera, data):
        """

        :param camera:
        :param data:
        :return:
        """
        ws = self.__wb.get_sheet_by_name('几何失真、帧频率与工频干扰')
        if 'front' == camera:
            ws['D51'] = data[0]  # 50hz 800lux
            ws['E51'] = data[1]  # 50hz 200lux
            ws['F51'] = data[2]  # 50hz 25lux
            ws['D52'] = data[3]  # 60hz 800lux
            ws['E52'] = data[4]  # 60hz 200lux
            ws['F52'] = data[5]  # 60hz 25lux
        else:  # for main camera
            ws['G51'] = data[0]  # 50hz 800lux
            ws['H51'] = data[1]  # 50hz 200lux
            ws['I51'] = data[2]  # 50hz 25lux
            ws['G52'] = data[3]  # 60hz 800lux
            ws['H52'] = data[4]  # 60hz 200lux
            ws['I52'] = data[5]  # 60hz 25lux

    def write_color_uniform_data(self, camera, data, sub_item):
        """

        :param sub_item:
        :param camera:
        :param data:
        :return:
        """
        ws = self.__wb.get_sheet_by_name('像面色彩均匀度')
        row_idx = 0
        col_idx = 0

        row_start_idx = 23
        col_start_idx = 5   # E
        row_idx = 0
        if 'main' == camera:
            row_start_idx = 53

        if 'D65' == sub_item:
            row_idx = row_start_idx + 9
        elif 'TL84' == sub_item:
            row_idx = row_start_idx + 18
        else:  # A
            row_idx = row_start_idx

        for i in range(9):
            ws.cell(row=row_idx+2, column=col_start_idx+i, value=data[0][i])  # B
            ws.cell(row=row_idx+1, column=col_start_idx+i, value=data[1][i])  # G
            ws.cell(row=row_idx, column=col_start_idx+i, value=data[2][i])  # R

    def write_luma_uniform_data(self, camera, data):
        """

        :param camera:
        :param data:
        :return:
        """
        ws = self.__wb.get_sheet_by_name('像面亮度均匀度')
        if 'front' == camera:
            ws['D15'] = data[0]
            ws['E15'] = data[1]
            ws['F15'] = data[2]
            ws['G15'] = data[3]
        else:  # for main camera
            ws['I15'] = data[0]
            ws['J15'] = data[1]
            ws['K15'] = data[2]
            ws['L15'] = data[3]

    def write_color_accuracy_data(self, camera, data, sub_item):
        """

        :param sub_item:
        :param camera:
        :param data:
        :return:
        """
        ws = self.__wb.get_sheet_by_name('色彩还原误差与饱和度')
        row = 19
        col = 'E'  # Front camera
        if 'main' == camera:
            col = 'H'

        if 'D65' == sub_item:
            pass
        elif 'TL84' == sub_item:
            col = chr(ord(col) + 1)
        elif 'A' == sub_item:
            col = chr(ord(col) + 2)
        else:
            return

        for i in range(len(data)):
            cell = col + str(row + i)
            ws[cell] = data[i]

    def write_saturation_data(self, camera, data, sub_item):
        """

        :param camera:
        :param data:
        :param sub_item:
        :return:
        """
        ws = self.__wb.get_sheet_by_name('色彩还原误差与饱和度')
        row = 54  # front camera
        if 'main' == camera:
            row = 55

        if 'D65' == sub_item:
            col = 'F'
        elif 'TL84' == sub_item:
            col = 'G'
        elif 'CWF' == sub_item:
            col = 'H'
        elif 'A' == sub_item:
            col = 'I'
        elif 'H' == sub_item:
            col = 'J'
        else:
            return

        cell = col + str(row)
        ws[cell] = data

    def write_wb_data(self, camera, data, sub_item):
        """

        :param camera:
        :param data:
        :param sub_item:
        :return:
        """
        ws = self.__wb.get_sheet_by_name('白平衡')
        col = 'F'  # front camera
        if 'main' == camera:
            col = 'J'

        if 'D65' == sub_item:
            row = 15
        elif 'TL84' == sub_item:
            row = 19
        elif 'CWF' == sub_item:
            row = 23
        elif 'A' == sub_item:
            row = 27
        else:
            return

        for i in range(len(data)):
            cell = col + str(row + i)
            ws[cell] = data[i]

    def write_free_data(self, camera, data, sub_item):
        """

        :param camera:
        :param data:  [file_name, awb, saturation, color_accuracy]
        :param sub_item: index
        :return:
        """
        if 0 == sub_item:
            self.__ws = self.__wb.create_sheet('Free ColorChecker', -1)
            align = Alignment(horizontal='center', vertical='center')
            font = Font(name=u'宋体', bold=True)
            border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
            cells = ['A1', 'B1', 'F1', 'G1']
            for cell in cells:
                self.__ws[cell].font = font
                self.__ws[cell].alignment = align
                self.__ws[cell].border = border
            self.__ws['A1'] = '文件名'
            self.__ws.merge_cells('B1:E1')
            self.__ws['B1'] = '白平衡（20-23）'
            self.__ws['F1'] = '饱和度'
            self.__ws.merge_cells('G1:AD1')
            self.__ws['G1'] = '色彩还原'

        if self.__ws == '':
            return
        row_start = 2
        self.__ws.cell(row=sub_item+row_start, column=1).value = data[0]  # file name
        for i in range(len(data[1])):
            self.__ws.cell(row=sub_item + row_start, column=2+i).value = data[1][i]  # awb
        self.__ws.cell(row=sub_item + row_start, column=6).value = str(round(data[2] * 100, 2)) + '%'  # saturation
        for i in range(len(data[3])):
            self.__ws.cell(row=sub_item + row_start, column=7 + i).value = data[3][i]  # color accuracy

    def write_corner_data(self, camera, data, sub_item):
        """

        :param camera:
        :param data: [module, light condition, max value]
        :param sub_item: corner module index
        :return:
        """
        ws = self.__wb.get_sheet_by_name('白平衡')

        row = 37 + sub_item * 4
        cell_module = 'C' + str(row)
        col = 'F'  # Front camera
        if 'main' == camera:
            col = 'J'

        if 'D65' == data[1]:
            row += 0
        elif 'TL84' == data[1]:
            row += 1
        elif 'CWF' == data[1]:
            row += 2
        elif 'A' == data[1]:
            row += 3
        else:
            return

        cell = col + str(row)
        ws[cell] = data[2]
        ws[cell_module] = data[0]

    def write_dr_data(self, camera, data):
        """

        :param camera:
        :param data:
        :return:
        """
        ws = self.__wb.get_sheet_by_name('动态范围')
        if 'front' == camera:
            ws['D51'] = data[0]  # 50hz 800lux
        else:  # for main camera
            ws['G51'] = data[0]  # 50hz 800lux